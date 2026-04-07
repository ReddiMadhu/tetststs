"""
output_builder.py — Assemble the final XLSX or CSV output.

The column order is driven by AIR_OUTPUT_COLUMNS / RMS_OUTPUT_COLUMNS constants.
Unmapped source columns are appended alphabetically at the end.
A second sheet "QA_Summary" provides a quality report.
"""
import csv
import io
import logging
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger("output_builder")

# ── AIR output column order ────────────────────────────────────────────────────
# Follows the AIR Touchstone / Common User Schema (CUC) format.

AIR_OUTPUT_COLUMNS: List[str] = [
    # Identity
    "PolicyID", "InsuredName", "LocationID", "LocationName",
    # Address
    "Street", "City", "Area", "PostalCode", "CountryISO",
    # Coordinates
    "Latitude", "Longitude",
    # Construction & Occupancy
    "ConstructionCodeType", "ConstructionCode",
    "OccupancyCodeType", "OccupancyCode",
    # Building attributes
    "RiskCount", "NumberOfStories", "GrossArea",
    "YearBuilt", "YearRetrofitted",
    # Secondary modifiers
    "SprinklerSystem", "RoofGeometry", "FoundationType",
    "WallSiding", "SoftStory", "WallType",
    # Financials
    "BuildingValue", "ContentsValue", "TimeElementValue",
    "Currency", "LineOfBusiness",
]

# ── RMS output column order ────────────────────────────────────────────────────
# Follows the RMS Exposure Data Module (EDM) standard.

RMS_OUTPUT_COLUMNS: List[str] = [
    # Core identity & geo
    "ACCNTNUM", "LOCNUM", "LOCNAME",
    "STREETNAME", "CITY", "STATECODE", "POSTALCODE", "CNTRYCODE",
    "Latitude", "Longitude",
    # Structural & secondary modifiers
    "BLDGSCHEME", "BLDGCLASS", "OCCSCHEME", "OCCTYPE",
    "NUMBLDGS", "NUMSTORIES", "FLOORAREA",
    "YEARBUILT", "YEARUPGRAD",
    "SPRINKLER", "ROOFGEOM", "FOUNDATION", "CLADDING", "SOFTSTORY", "WALLTYPE",
    # Values (peril-specific)
    "EQCV1VAL", "EQCV2VAL", "EQCV3VAL",
    "WSCV1VAL", "WSCV2VAL", "WSCV3VAL",
    "TOCV1VAL", "TOCV2VAL", "TOCV3VAL",
    "FLCV1VAL", "FLCV2VAL", "FLCV3VAL",
    "TRCV1VAL", "TRCV2VAL", "TRCV3VAL",
    "FRCV1VAL", "FRCV2VAL", "FRCV3VAL",
    # Currency (peril-specific)
    "EQCV1LCUR", "EQCV2LCUR", "EQCV3LCUR",
    "WSCV1LCUR", "WSCV2LCUR", "WSCV3LCUR",
    "TOCV1LCUR", "TOCV2LCUR", "TOCV3LCUR",
    "FLCV1LCUR", "FLCV2LCUR", "FLCV3LCUR",
    "TRCV1LCUR", "TRCV2LCUR", "TRCV3LCUR",
    "FRCV1LCUR", "FRCV2LCUR", "FRCV3LCUR",
]

# ── Styling constants ──────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
QA_HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
QA_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
ALT_ROW_FILL = PatternFill("solid", fgColor="EBF3FB")


# ── Build XLSX ─────────────────────────────────────────────────────────────────

def build_xlsx(
    rows: List[Dict[str, Any]],
    unmapped_cols: List[str],
    flags: List[Dict],
    target_format: str = "AIR",
    upload_id: str = "",
) -> io.BytesIO:
    """
    Build a streaming XLSX workbook with:
      Sheet 1 — Locations (all processed rows in canonical column order)
      Sheet 2 — QA_Summary (stats + flag details)
    Returns a BytesIO buffer positioned at 0.
    """
    wb = Workbook(write_only=True)

    # ── Sheet 1: Locations ─────────────────────────────────────────────────────
    ws_loc = wb.create_sheet("Locations")

    base_cols = AIR_OUTPUT_COLUMNS if target_format == "AIR" else RMS_OUTPUT_COLUMNS
    # Only output the canonical schema columns — unmapped source columns are
    # excluded from the Locations sheet to keep output clean and EDM-compliant.
    final_cols = base_cols

    # Set AutoFit column widths based on the minimum length of standard column sizes + buffer
    for i, col_name in enumerate(final_cols, start=1):
        col_letter = get_column_letter(i)
        # Ensure column width is wide enough for the header text at least
        ws_loc.column_dimensions[col_letter].width = max(len(col_name) + 2, 12)

    # Styled header row
    header_cells = []
    for col_name in final_cols:
        from openpyxl.cell.cell import WriteOnlyCell
        cell = WriteOnlyCell(ws_loc, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        header_cells.append(cell)
    ws_loc.append(header_cells)

    if target_format == "RMS":
        _clone_rms_perils(rows)
        _format_rms_years(rows)

    # Data rows
    for i, row in enumerate(rows):
        row_data = []
        for col_name in final_cols:
            val = row.get(col_name)
            # Coerce None-like values
            if val is not None:
                if isinstance(val, float) and (val != val):  # NaN
                    val = None
            row_data.append(val)

        data_cells = []
        for val in row_data:
            from openpyxl.cell.cell import WriteOnlyCell
            cell = WriteOnlyCell(ws_loc, value=val)
            if i % 2 == 1:
                cell.fill = ALT_ROW_FILL
            data_cells.append(cell)
        ws_loc.append(data_cells)

    # ── Sheet 2: QA_Summary ────────────────────────────────────────────────────
    ws_qa = wb.create_sheet("QA_Summary")

    stats = _compute_qa_stats(rows, flags, target_format, upload_id)
    from openpyxl.cell.cell import WriteOnlyCell

    # Section header
    def qa_header_cell(text: str):
        c = WriteOnlyCell(ws_qa, value=text)
        c.font = QA_HEADER_FONT
        c.fill = QA_HEADER_FILL
        return c

    def qa_cell(text: Any):
        return WriteOnlyCell(ws_qa, value=text)

    ws_qa.append([qa_header_cell("Field"), qa_header_cell("Value")])
    for label, value in stats["summary_rows"]:
        ws_qa.append([qa_cell(label), qa_cell(value)])

    ws_qa.append([])  # blank row

    if flags:
        ws_qa.append([qa_header_cell("Row"), qa_header_cell("Field"),
                      qa_header_cell("Issue"), qa_header_cell("Message")])
        for flag in flags:
            ws_qa.append([
                qa_cell(flag.get("row_index", "")),
                qa_cell(flag.get("field", "")),
                qa_cell(flag.get("issue", "")),
                qa_cell(flag.get("message", "")),
            ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_csv(
    rows: List[Dict[str, Any]],
    unmapped_cols: List[str],
    target_format: str = "AIR",
) -> io.BytesIO:
    """Build a CSV output in the canonical column order. Returns BytesIO."""
    base_cols = AIR_OUTPUT_COLUMNS if target_format == "AIR" else RMS_OUTPUT_COLUMNS
    # Only output the canonical schema columns
    final_cols = base_cols

    if target_format == "RMS":
        _clone_rms_perils(rows)
        _format_rms_years(rows)

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=final_cols, extrasaction="ignore",
                             lineterminator="\n")
    writer.writeheader()
    for row in rows:
        writer.writerow({col: row.get(col, "") for col in final_cols})

    raw = buf.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility
    return io.BytesIO(raw)




def _clone_rms_perils(rows: List[Dict]) -> None:
    """Clone EQ values to WS, TO, FL, TR, and FR peril columns."""
    for row in rows:
        # Values
        val1 = row.get("EQCV1VAL")
        val2 = row.get("EQCV2VAL")
        val3 = row.get("EQCV3VAL")
        
        # Currencies
        cur1 = row.get("EQCV1LCUR")
        cur2 = row.get("EQCV2LCUR")
        cur3 = row.get("EQCV3LCUR")

        for prefix in ("WSCV", "TOCV", "FLCV", "TRCV", "FRCV"):
            if val1 is not None and row.get(f"{prefix}1VAL") is None: row[f"{prefix}1VAL"] = val1
            if val2 is not None and row.get(f"{prefix}2VAL") is None: row[f"{prefix}2VAL"] = val2
            if val3 is not None and row.get(f"{prefix}3VAL") is None: row[f"{prefix}3VAL"] = val3
            
            if cur1 is not None and row.get(f"{prefix}1LCUR") is None: row[f"{prefix}1LCUR"] = cur1
            if cur2 is not None and row.get(f"{prefix}2LCUR") is None: row[f"{prefix}2LCUR"] = cur2
            if cur3 is not None and row.get(f"{prefix}3LCUR") is None: row[f"{prefix}3LCUR"] = cur3
            
def _format_rms_years(rows: List[Dict]) -> None:
    """Format YEARBUILT and YEARUPGRAD to MM/DD/YYYY standard for RMS EDM."""
    for row in rows:
        for col in ("YEARBUILT", "YEARUPGRAD"):
            val = row.get(col)
            if val is not None and val != "":
                try:
                    y = int(val)
                    if y == 9999:
                        row[col] = "31//12/9999"
                    elif 1700 <= y <= 2026:
                        row[col] = f"01/01/{y}"
                except (ValueError, TypeError):
                    pass

# ── QA Stats ───────────────────────────────────────────────────────────────────

def _compute_qa_stats(
    rows: List[Dict],
    flags: List[Dict],
    target_format: str,
    upload_id: str,
) -> Dict:
    total = len(rows)
    geo_failed = sum(1 for r in rows if r.get("GeocodingStatus") not in ("OK", "PROVIDED"))
    low_occ = sum(1 for r in rows
                  if (r.get("Occupancy_Confidence") or 1.0) < 0.70)
    low_const = sum(1 for r in rows
                    if (r.get("Construction_Confidence") or 1.0) < 0.70)
    currency_conflicts = sum(1 for r in rows if r.get("Currency_Conflicts"))
    bad_years = sum(1 for r in rows if r.get("Year_Built_Flag") not in (None, "VALID"))
    review_rows = len({f["row_index"] for f in flags})

    fmt_label = "AIR Touchstone XLSX" if target_format == "AIR" else "RMS EDM CSV"
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    summary_rows = [
        ("Total rows processed", total),
        ("Rows with geocoding failures", geo_failed),
        ("Rows with low occupancy confidence (<0.70)", low_occ),
        ("Rows with low construction confidence (<0.70)", low_const),
        ("Rows with currency conflicts", currency_conflicts),
        ("Rows with implausible / out-of-range years", bad_years),
        ("Rows requiring manual review (any flag)", review_rows),
        ("Total flags", len(flags)),
        ("Target format", fmt_label),
        ("Upload ID", upload_id),
        ("Processing timestamp", ts),
    ]
    return {"summary_rows": summary_rows}
