"""
main.py â€” FastAPI application for the CAT Modeling Data Pipeline.
"""
import io
import json
import logging
import os
import pathlib
import asyncio
from contextlib import asynccontextmanager
from typing import Any, Dict, List, Optional

import pandas as pd
from dotenv import load_dotenv
from fastapi import FastAPI, File, Form, HTTPException, Query, UploadFile, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sse_starlette.sse import EventSourceResponse

# â”€â”€ Load env early â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
load_dotenv()

# â”€â”€ Local imports â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
import session as session_store
from rules import BusinessRulesConfig
from models import (
    UploadResponse, SuggestColumnsResponse, ColumnSuggestion,
    ConfirmColumnsRequest, ConfirmColumnsResponse,
    GeocodeResponse, MapCodesResponse, NormalizeResponse,
    ReviewResponse, FlagEntry, CorrectRequest, CorrectResponse,
    SessionInfoResponse,
)
from output_builder import (
    build_xlsx, build_tsv, build_account_xlsx, build_account_tsv,
    generate_location_data, _get_account_rows
)

# Agents
import agents.geocoder as geocoder
import agents.cat.code_mapper as code_mapper
import agents.cat.mapping_memory as mapping_memory
from agents.cat.column_mapper import suggest_columns, validate_required_fields
from agents.normalizer import normalize_all_rows

logging.basicConfig(
    level=os.getenv("LOG_LEVEL", "INFO"),
    format="%(asctime)s [%(name)s] %(levelname)s â€” %(message)s",
)
logger = logging.getLogger("main")

# ── LLM Summary Generator (LangChain ChatOpenAI) ──────────────────────────────
from langchain_openai import ChatOpenAI
from langchain_core.messages import SystemMessage, HumanMessage

_SUMMARY_LLM: Optional[ChatOpenAI] = None

_SUMMARY_SYSTEM = (
    "You are a senior insurance data analyst. Given pipeline transformation stats, "
    "produce a concise executive summary as a list of key points. "
    "Each point should be ONE short sentence covering a specific insight. "
    "Aim for 4-6 bullet points. Be dense with information. "
    "Mention specific numbers: rows processed, codes mapped, methods used, flags raised. "
    "Sound professional and confident about CAT modeling / underwriting terminology. "
    "Do NOT invent numbers, only reference data from the provided context. "
    "Frame flags constructively (e.g. flagged for expert review). "
    'Return ONLY a JSON object: {"points": ["Point 1 here.", "Point 2 here.", ...]}'
)


def _init_summary_model() -> None:
    """Initialize the ChatOpenAI model for pipeline summary generation."""
    global _SUMMARY_LLM
    api_key = os.getenv("OPENAI_API_KEY", "")
    base_url = os.getenv("OPENAI_BASE_URL", None)
    model = os.getenv("OPENAI_MODEL", "gpt-4o-mini")

    if not api_key:
        logger.warning("No OPENAI_API_KEY found, LLM summaries disabled.")
        return

    kwargs = {
        "model": model,
        "api_key": api_key,
        "temperature": 0.7,
    }
    if base_url:
        kwargs["base_url"] = base_url

    _SUMMARY_LLM = ChatOpenAI(**kwargs)
    logger.info(f"Summary LLM initialized (ChatOpenAI model={model}, base_url={base_url or 'default'}).")


def _generate_llm_summary(step_name: str, stats_context: dict, fallback_text: str) -> str:
    """
    Call ChatOpenAI to generate a list of key-point summaries from pipeline stats.
    Returns a JSON string containing the points array, or fallback_text on failure.
    """
    if _SUMMARY_LLM is None:
        return fallback_text

    step_label = ("Occupancy & Construction Code Mapping"
                  if step_name == "code_mapping"
                  else "Value Normalization")
    user_prompt = (
        f"Summarize the {step_label} step.\n\n"
        f"STATS:\n{json.dumps(stats_context, indent=2, default=str)}\n\n"
        f"Write 4-6 key insight bullet points."
    )

    try:
        messages = [
            SystemMessage(content=_SUMMARY_SYSTEM),
            HumanMessage(content=user_prompt),
        ]
        response = _SUMMARY_LLM.invoke(messages)
        raw_text = response.content.strip()

        # Strip potential markdown formatting (```json ... ```)
        if raw_text.startswith("```json"):
            raw_text = raw_text.replace("```json", "", 1)
        if raw_text.startswith("```"):
            raw_text = raw_text.replace("```", "", 1)
        if raw_text.endswith("```"):
            raw_text = raw_text[::-1].replace("```", "", 1)[::-1]
        raw_text = raw_text.strip()

        logger.info(f"LLM Raw Response: {raw_text}")

        parsed = json.loads(raw_text)

        # Support both 'points' array and legacy 'summary' string
        points = parsed.get("points", [])
        if points and isinstance(points, list):
            logger.info(f"LLM summary generated for {step_name} ({len(points)} points).")
            return json.dumps({"points": points})

        summary = parsed.get("summary", "").strip()
        if summary:
            logger.info(f"LLM returned legacy summary for {step_name}, converting to points.")
            sentences = [s.strip() for s in summary.replace(". ", ".\n").split("\n") if s.strip()]
            return json.dumps({"points": sentences})

        logger.warning(f"LLM returned JSON without 'points' or 'summary' key: {parsed}")
        return fallback_text
    except Exception as e:
        logger.error(f"LLM summary generation FAILED for {step_name}: {type(e).__name__}: {e}", exc_info=True)
        return fallback_text

# Global event queues for SSE
_event_queues: Dict[str, List[asyncio.Queue]] = {}

def dispatch_event(upload_id: str, agent_id: str, event: str, message: str = "", result: dict = None):
    """Utility to push an event to all connected SSE clients for a session."""
    if upload_id in _event_queues:
        data = {
            "agent_id": agent_id,
            "event": event,
            "message": message,
            "result": result
        }
        for q in _event_queues[upload_id]:
            q.put_nowait(data)

# â”€â”€ Lifespan: pre-build TF-IDF and 8RN6JojL-pjXIGh76Oa5QL0yyx5c-j start TTL vOOhXGfgaw4V3PQ cleanup â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@asynccontextmanager
async def lifespan(app: FastAPI):
    logger.info("Starting CAT pipeline â€” loading TF-IDF indexesâ€¦")
    code_mapper.build_tfidf_indexes()
    geocoder.load_reference_data()
    _init_summary_model()
    session_store.start_ttl_cleanup()
    logger.info("Startup complete.")
    yield
    logger.info("Shutting down.")


# â”€â”€ App â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

app = FastAPI(
    title="CAT Modeling Data Pipeline",
    description="Upload, map, geocode, classify, and normalise property exposure data for AIR/RMS CAT models.",
    version="1.0.0",
    lifespan=lifespan,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Ontology & Rules API ─────────────────────────────────────────────────────────
from ontology_router import router as ontology_api_router
app.include_router(ontology_api_router, prefix="/api")

# â”€â”€ Helpers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def _get_session_or_404(upload_id: str) -> dict:
    try:
        return session_store.require_session(upload_id)
    except KeyError:
        raise HTTPException(status_code=404, detail=f"Session '{upload_id}' not found or expired.")


def _require_stage(session: dict, stage: str, endpoint: str) -> None:
    if not session.get("stages_complete", {}).get(stage):
        raise HTTPException(
            status_code=422,
            detail=f"Stage '{stage}' must be completed before calling '{endpoint}'.",
        )


def _load_iso4217() -> set:
    p = pathlib.Path(__file__).parent / "reference" / "iso4217_currency.json"
    if p.exists():
        data = json.loads(p.read_text(encoding="utf-8"))
        return set(data.keys())
    return set()


_VALID_CURRENCIES = _load_iso4217()


def _enrich_excel_formats(content: bytes, df: pd.DataFrame) -> pd.DataFrame:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        col_currencies = {}
        for col_idx in range(1, ws.max_column + 1):
            if col_idx - 1 >= len(df.columns): break
            col_name = df.columns[col_idx - 1]
            for r in range(2, min(ws.max_row + 1, 10)):
                cell = ws.cell(row=r, column=col_idx)
                if cell.value is not None:
                    fmt = str(cell.number_format or "")
                    for sym in ["$", "â‚¬", "Â£", "Â¥", "â‚¹"]:
                        if sym in fmt:
                            col_currencies[col_name] = sym
                            break
                    if col_name in col_currencies:
                        break
        for col, sym in col_currencies.items():
            df[col] = df[col].apply(lambda v: f"{sym}{v}" if pd.notnull(v) and str(v).strip() != "" else v)
    except Exception as e:
        logger.warning(f"Could not extract Excel currency formats: {e}")
    return df


# â”€â”€ Step 1: Upload â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.post("/upload", response_model=UploadResponse, tags=["Pipeline"])
async def upload(
    file: UploadFile = File(...),
    target_format: str = Query("AIR", pattern="^(AIR|RMS)$"),
    rules_config: Optional[str] = Form(None),
):
    content = await file.read()
    fname = (file.filename or "").lower()

    try:
        rules_dict = json.loads(rules_config) if rules_config and rules_config.strip() else {}
        rules = BusinessRulesConfig(**rules_dict)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Invalid rules_config: {exc}")

    try:
        if fname.endswith(".csv"):
            try:
                df = pd.read_csv(io.BytesIO(content), dtype=str, keep_default_na=False)
            except UnicodeDecodeError:
                df = pd.read_csv(io.BytesIO(content), dtype=str, keep_default_na=False, encoding="latin-1")
        elif fname.endswith((".xlsx", ".xls")):
            df = pd.read_excel(io.BytesIO(content), dtype=str, keep_default_na=False, sheet_name=0)
            df = _enrich_excel_formats(content, df)
        else:
            raise HTTPException(status_code=400, detail="Unsupported file type. Upload CSV or XLSX.")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"File parsing error: {exc}")

    if df.empty:
        raise HTTPException(status_code=400, detail="File contains no data rows.")

    seen: Dict[str, int] = {}
    new_cols = []
    for col in df.columns:
        col = str(col).strip()
        if col in seen:
            seen[col] += 1
            new_cols.append(f"{col}_{seen[col]}")
        else:
            seen[col] = 1
            new_cols.append(col)
    df.columns = new_cols

    df = df.map(lambda v: v.strip() if isinstance(v, str) and v.strip() != "" else None)

    headers = list(df.columns)
    raw_rows = df.to_dict(orient="records")
    # Provide the full dataset as the sample to display in the frontend immediately
    sample = raw_rows

    upload_id = session_store.create_session({
        "target_format": target_format,
        "rules_config": rules.model_dump(),
        "raw_rows": raw_rows,
        "original_raw_rows": raw_rows,   # immutable snapshot for diffs
        "headers": headers,
        "sample": sample,
        "column_map": {},
        "unmapped_cols": [],
        "geo_rows": [],
        "code_map": {},
        "final_rows": [],
        "stages_complete": {"upload": True},
    })

    return UploadResponse(
        upload_id=upload_id,
        row_count=len(raw_rows),
        headers=headers,
        sample=sample,
        target_format=target_format,
    )


# â”€â”€ Internal helpers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def _apply_column_map(raw_rows: List[Dict], column_map: Dict[str, Optional[str]]) -> List[Dict]:
    """
    Rename source columns to canonical names while PRESERVING all other columns.
    Columns that are mapped get renamed (src_col key removed, canonical key added).
    Columns with no mapping are kept under their original key.
    """
    canonical_claimed_by: Dict[str, str] = {}  # canonical -> src_col
    for src_col, canonical in column_map.items():
        if canonical is None:
            continue
        if canonical not in canonical_claimed_by:
            canonical_claimed_by[canonical] = src_col
    # Build reverse: src_col -> canonical (so we know which src keys to drop)
    src_to_canonical = {v: k for k, v in canonical_claimed_by.items()}

    result = []
    for row in raw_rows:
        new_row: Dict[str, Any] = dict(row)          # start with ALL existing columns
        for src_col, canonical in src_to_canonical.items():
            if src_col in new_row:
                if canonical not in new_row or new_row[canonical] is None:
                    new_row[canonical] = new_row[src_col]  # write under canonical key
                if src_col != canonical:                   # remove old key only if different
                    del new_row[src_col]
        result.append(new_row)
    return result


def _find_code_columns(column_map: Dict, field_type: str, target: str):
    if field_type == "occupancy":
        scheme_options = ["OccupancyCodeType"] if target == "AIR" else ["OCCSCHEME"]
        value_options  = ["OccupancyCode"] if target == "AIR" else ["OCCTYPE"]
    else:
        scheme_options = ["ConstructionCodeType"] if target == "AIR" else ["BLDGSCHEME"]
        value_options  = ["ConstructionCode"]  if target == "AIR" else ["BLDGCLASS"]

    canonical_to_src = {v: k for k, v in column_map.items() if v}
    scheme_col = next((canonical_to_src[c] for c in scheme_options if c in canonical_to_src), "")
    value_col  = next((canonical_to_src[c] for c in value_options  if c in canonical_to_src), "")
    return scheme_col, value_col


# â”€â”€ SSE Streaming Endpoint â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.get("/stream/{upload_id}", tags=["Event Stream"])
async def message_stream(upload_id: str, request: Request):
    """SSE endpoint for agent event propagation to the frontend."""
    _get_session_or_404(upload_id) # Verify session exists

    q = asyncio.Queue()
    if upload_id not in _event_queues:
        _event_queues[upload_id] = []
    _event_queues[upload_id].append(q)

    async def event_generator():
        try:
            while True:
                if await request.is_disconnected():
                    break
                data = await q.get()
                yield {"data": json.dumps(data)}
        finally:
            if upload_id in _event_queues:
                if q in _event_queues[upload_id]:
                    _event_queues[upload_id].remove(q)
                if not _event_queues[upload_id]:
                    del _event_queues[upload_id]

    return EventSourceResponse(event_generator())


# â”€â”€ Step 2: Address Normalization â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€


# â”€â”€ Step 2a: Suggest columns â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.get("/session/{upload_id}/status", tags=["Pipeline"])
def session_status(upload_id: str):
    """Return pipeline stage completion flags for frontend hydration."""
    session = _get_session_or_404(upload_id)
    return {
        "upload_id": upload_id,
        "target_format": session.get("target_format", "AIR"),
        "stages_complete": session.get("stages_complete", {}),
        "row_count": len(session.get("raw_rows", [])),
        "headers": session.get("headers", []),
        "sample": session.get("sample", []),
    }


@app.get("/suggest-columns/{upload_id}", response_model=SuggestColumnsResponse, tags=["Pipeline"])
def suggest_columns_endpoint(upload_id: str, target_format: Optional[str] = Query(None, pattern="^(AIR|RMS)$")):
    session = _get_session_or_404(upload_id)
    
    if target_format and target_format != session.get("target_format"):
        session["target_format"] = target_format
        session_store.update_session(upload_id, {"target_format": target_format})

    # Use geo_rows headers post-geocoding so the mapper sees geocoded columns
    geo_rows = session.get("geo_rows", [])
    if geo_rows:
        raw_rows = geo_rows
        headers = list(geo_rows[0].keys()) if geo_rows else session["headers"]
    else:
        raw_rows = session["raw_rows"]
        headers = session["headers"]

    sample_values: Dict[str, List[Any]] = {}
    for col in headers:
        vals = [r[col] for r in raw_rows[:20] if r.get(col) is not None][:3]
        sample_values[col] = vals

    result = suggest_columns(
        source_columns=headers,
        sample_values=sample_values,
        target_format=session["target_format"],
        fuzzy_threshold=session["rules_config"].get("fuzzy_llm_fallback_threshold", 72),
        cutoff=session["rules_config"].get("fuzzy_score_cutoff", 50),
    )

    suggestions_typed = {
        col: [ColumnSuggestion(**s) for s in sug_list]
        for col, sug_list in result["suggestions"].items()
    }
    return SuggestColumnsResponse(
        suggestions=suggestions_typed,
        unmapped=result["unmapped"],
        memory_count=result.get("memory_count", 0),
        sample_values=sample_values,
    )


# â”€â”€ Step 2b: Confirm columns â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.post("/confirm-columns/{upload_id}", response_model=ConfirmColumnsResponse, tags=["Pipeline"])
def confirm_columns(upload_id: str, body: ConfirmColumnsRequest):
    session = _get_session_or_404(upload_id)

    column_map = body.column_map
    canonical_to_sources: Dict[str, List[str]] = {}
    for src_col, canonical in column_map.items():
        if canonical is None: continue
        canonical_to_sources.setdefault(canonical, []).append(src_col)

    duplicate_violations = { c: s for c, s in canonical_to_sources.items() if len(s) > 1 }

    if duplicate_violations:
        details = "; ".join(f"'{c}' claimed by: {s}" for c, s in duplicate_violations.items())
        raise HTTPException(status_code=422, detail=f"Mapping violation: {details}")

    unmapped = [k for k, v in column_map.items() if v is None]
    mapped_count = len(column_map) - len(unmapped)

    missing_required = validate_required_fields(column_map, session["target_format"])
    warnings = [f"Required field '{f}' is not mapped." for f in missing_required]

    session_store.update_session(upload_id, {
        "column_map": column_map,
        "unmapped_cols": unmapped,
    })
    session_store.session_mark_stage(upload_id, "column_map")

    try:
        mapping_memory.record_confirmed(column_map, session["target_format"])
    except Exception as exc:
        pass

    return ConfirmColumnsResponse(
        upload_id=upload_id,
        mapped_count=mapped_count,
        unmapped_cols=unmapped,
        missing_required=missing_required,
        warnings=warnings,
    )


# â”€â”€ Step 3: Geocode â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.post("/geocode/{upload_id}", response_model=GeocodeResponse, tags=["Pipeline"])
async def geocode_endpoint(upload_id: str):
    session = _get_session_or_404(upload_id)
    # Geocoding in this flow does not require column mapping first.
    
    rules_config = BusinessRulesConfig(**session["rules_config"])
    raw_rows = session["raw_rows"]

    agent_id = "geocoder"
    dispatch_event(upload_id, agent_id, "start")
    dispatch_event(upload_id, agent_id, "log", f"Started geocoding {len(raw_rows)} rows via Geoapify API...")
    
    # Call geocoder
    geocoded_count = 0
    provided_count = 0
    failed_count = 0
    new_flags: List[dict] = []
    geo_rows: List[dict] = []
    
    target_format = session.get("target_format", "AIR")
    
    for idx, row in enumerate(raw_rows):
        if idx % 10 == 0:
            dispatch_event(upload_id, agent_id, "log", f"Processing {idx}/{len(raw_rows)} rows...")

        # In geocoder.py from catai, it expects a column map to find the address fields.
        # But wait! We do Address Normalization first in this MVP, which outputs standardized fields.
        # Let's pass a dummy map since Address Normalization standardized the fields.
        dummy_col_map = {
            "FullAddress": "FullAddress",
            "Street": "Street", "City": "City", "Area": "Area", 
            "PostalCode": "PostalCode", "CountryISO": "CountryISO",
            "STREETNAME": "STREETNAME", "STATECODE": "STATECODE",
            "CNTRYCODE": "CNTRYCODE", "CITY": "CITY"
        }
        
        geo_fields = geocoder.process_row_geocoding(row, dummy_col_map, target_format=target_format)
        row.update(geo_fields)
        geo_rows.append(row)

        status = geo_fields.get("GeocodingStatus", "")
        source = geo_fields.get("Geosource", "")

        if source == "Provided": provided_count += 1
        elif status == "OK": geocoded_count += 1
        else: failed_count += 1

    # Update headers to include any new geocoded columns
    existing_headers = session.get("headers", [])
    existing_header_set = set(existing_headers)
    new_geo_cols = [k for k in (geo_rows[0].keys() if geo_rows else []) if k not in existing_header_set]
    updated_headers = existing_headers + new_geo_cols

    session_store.update_session(upload_id, {"geo_rows": geo_rows, "headers": updated_headers})
    session_store.append_flags(upload_id, new_flags)
    session_store.session_mark_stage(upload_id, "geocoding")
    dispatch_event(upload_id, "pipeline", "stage_complete", "geocoding")

    result = {
        "geocoded": geocoded_count,
        "provided": provided_count,
        "failed": failed_count,
        "flags_added": len(new_flags)
    }
    
    dispatch_event(upload_id, agent_id, "done", result=result)

    diff_data = _get_session_diff_data(session, "geocode")

    return GeocodeResponse(
        upload_id=upload_id, 
        total_rows=len(raw_rows), 
        sample=geo_rows[:10],
        headers=session.get("headers", []),
        diff_data=diff_data,
        **result
    )


# â”€â”€ Step 4: Map codes â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.post("/map-codes/{upload_id}", response_model=MapCodesResponse, tags=["Pipeline"])
async def map_codes_endpoint(upload_id: str):
    session = _get_session_or_404(upload_id)
    _require_stage(session, "geocoding", "/map-codes")
    _require_stage(session, "column_map", "/map-codes")

    geo_rows = session["geo_rows"]
    column_map = session["column_map"]
    rules_config = BusinessRulesConfig(**session["rules_config"])
    target = session["target_format"]
    
    agent_id = "cat_code_mapper"
    dispatch_event(upload_id, agent_id, "start")

    occ_scheme_col, occ_value_col = _find_code_columns(column_map, "occupancy", target)
    const_scheme_col, const_value_col = _find_code_columns(column_map, "construction", target)

    target_occ_scheme = "OccupancyCodeType" if target == "AIR" else "OCCSCHEME"
    target_const_scheme = "ConstructionCodeType" if target == "AIR" else "BLDGSCHEME"

    new_flags: List[dict] = []
    enriched_rows = []
    code_map: Dict[str, Dict] = {}

    if occ_value_col:
        dispatch_event(upload_id, agent_id, "log", "Identifying unique occupancy combinations...")
        occ_items = code_mapper.extract_unique_pairs(geo_rows, occ_scheme_col, occ_value_col)
        occ_results = code_mapper.map_codes(occ_items, target, "occupancy", rules_config)
        for item in occ_items:
            key = code_mapper.build_row_key(item["scheme"], item["value"])
            result = occ_results.get(str(item["index"]), {})
            if result: code_map[f"occ|{key}"] = result

    if const_value_col:
        dispatch_event(upload_id, agent_id, "log", "Identifying unique construction combinations...")
        const_items = code_mapper.extract_unique_pairs(geo_rows, const_scheme_col, const_value_col)
        const_results = code_mapper.map_codes(const_items, target, "construction", rules_config)
        for item in const_items:
            key = code_mapper.build_row_key(item["scheme"], item["value"])
            result = const_results.get(str(item["index"]), {})
            if result: code_map[f"const|{key}"] = result

    occ_by_method: Dict[str, int] = {}
    const_by_method: Dict[str, int] = {}

    dispatch_event(upload_id, agent_id, "log", "Enriching rows with mapped canonical codes...")

    for idx, row in enumerate(geo_rows):
        row = dict(row)

        if occ_value_col:
            scheme = str(row.get(occ_scheme_col) or "").strip()
            value = str(row.get(occ_value_col) or "").strip()
            key = f"occ|{code_mapper.build_row_key(scheme, value)}"
            result = code_map.get(key, {})
            if result:
                row[occ_value_col] = result["code"]
                if not row.get(target_occ_scheme): row[target_occ_scheme] = "ATC" if target == "RMS" else "AIR"
                row["Occupancy_Code"]        = result["code"]
                row["Occupancy_Description"] = result["description"]
                row["Occupancy_Confidence"]  = result["confidence"]
                row["Occupancy_Method"]      = result["method"]
                row["Occupancy_Original"]    = result["original"]
                occ_by_method[result["method"]] = occ_by_method.get(result["method"], 0) + 1

        if const_value_col:
            scheme = str(row.get(const_scheme_col) or "").strip()
            value = str(row.get(const_value_col) or "").strip()
            key = f"const|{code_mapper.build_row_key(scheme, value)}"
            result = code_map.get(key, {})
            if result:
                row[const_value_col] = result["code"]
                scheme_override = result.get("scheme_override")
                if scheme_override == "ISF": row[target_const_scheme] = "ISF"
                elif target == "RMS" and scheme_override: row[target_const_scheme] = scheme_override
                elif not row.get(target_const_scheme): row[target_const_scheme] = target
                
                row["Construction_Code"]        = result["code"]
                row["Construction_Description"] = result["description"]
                row["Construction_Confidence"]  = result["confidence"]
                row["Construction_Method"]      = result["method"]
                row["Construction_Original"]    = result["original"]
                row["Construction_Scheme"]      = row.get(target_const_scheme, target)
                const_by_method[result["method"]] = const_by_method.get(result["method"], 0) + 1

        enriched_rows.append(row)

    session_store.update_session(upload_id, {
        "code_map": code_map,
        "final_rows": enriched_rows,
    })
    session_store.append_flags(upload_id, new_flags)
    session_store.session_mark_stage(upload_id, "code_mapping")
    # NOTE: stage_complete is NOT dispatched here — this is a merged endpoint.
    # The HTTP response (with diff_data) is the authoritative completion signal for the frontend.

    unique_occ = len([k for k in code_map if k.startswith("occ|")])
    unique_const = len([k for k in code_map if k.startswith("const|")])

    # ── Build AI summary ───────────────────────────────────────────────────
    total_rows = len(enriched_rows)
    total_occ_rows = sum(occ_by_method.values())
    total_const_rows = sum(const_by_method.values())

    # ── Call LLM for narrative summary ─────────────────────────────────────
    # Build stats context for the LLM
    occ_code_dist_ctx: Dict[str, int] = {}
    const_code_dist_ctx: Dict[str, int] = {}
    sample_mappings = []
    for k, v in code_map.items():
        code_val = v.get("code", "?")
        desc = v.get("description", "")
        label = f"{code_val} — {desc}" if desc else code_val
        if k.startswith("occ|"):
            occ_code_dist_ctx[label] = occ_code_dist_ctx.get(label, 0) + 1
        elif k.startswith("const|"):
            const_code_dist_ctx[label] = const_code_dist_ctx.get(label, 0) + 1
        if len(sample_mappings) < 6:
            sample_mappings.append({
                "type": "occupancy" if k.startswith("occ|") else "construction",
                "original": v.get("original", ""),
                "mapped_code": code_val,
                "mapped_desc": desc,
                "method": v.get("method", ""),
                "confidence": v.get("confidence", 0),
            })

    llm_stats_context = {
        "step": "code_mapping",
        "target_format": target,
        "total_rows": total_rows,
        "unique_occ_pairs": unique_occ,
        "unique_const_pairs": unique_const,
        "total_occ_classified": total_occ_rows,
        "total_const_classified": total_const_rows,
        "occ_by_method": occ_by_method,
        "const_by_method": const_by_method,
        "occ_code_distribution": dict(sorted(occ_code_dist_ctx.items(), key=lambda x: x[1], reverse=True)[:10]),
        "const_code_distribution": dict(sorted(const_code_dist_ctx.items(), key=lambda x: x[1], reverse=True)[:10]),
        "flags_count": len(new_flags),
        "sample_mappings": sample_mappings,
    }

    dispatch_event(upload_id, agent_id, "log", "Generating AI summary of mapping outcomes…")
    summary_text = _generate_llm_summary("code_mapping", llm_stats_context, "Occupancy and construction codes successfully mapped.")

    result = {
        "unique_occ_pairs": unique_occ,
        "unique_const_pairs": unique_const,
        "occ_by_method": occ_by_method,
        "const_by_method": const_by_method,
        "flags_added": len(new_flags),
    }

    dispatch_event(upload_id, agent_id, "done", result=result)

    diff_data = _get_session_diff_data(session, "map-codes")

    return MapCodesResponse(upload_id=upload_id, summary_text=summary_text, diff_data=diff_data, **result)


# â”€â”€ Step 5: Normalize â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.post("/normalize-values/{upload_id}", response_model=NormalizeResponse, tags=["Pipeline"])
async def normalize_endpoint(upload_id: str):
    session = _get_session_or_404(upload_id)
    _require_stage(session, "code_mapping", "/normalize")
    
    agent_id = "cat_normalizer"
    dispatch_event(upload_id, agent_id, "start")

    rules_config = BusinessRulesConfig(**session["rules_config"])
    # Apply column map to final_rows to ensure canonical names
    final_rows = session["final_rows"] or session["geo_rows"]
    column_map = session["column_map"]
    
    dispatch_event(upload_id, agent_id, "log", "Applying canonical column mappings...")
    final_rows_canonical = _apply_column_map(final_rows, column_map)

    valid_currencies = _VALID_CURRENCIES
    target_format = session.get("target_format", "AIR")

    dispatch_event(upload_id, agent_id, "log", "Normalizing years, floors, areas, and values...")
    normalized, new_flags = normalize_all_rows(final_rows_canonical, rules_config, valid_currencies, target_format=target_format)

    lob_col = "LineOfBusiness" if target_format == "AIR" else ""
    if lob_col and rules_config.line_of_business:
        for r in normalized:
            if not r.get(lob_col): r[lob_col] = rules_config.line_of_business

    session_store.update_session(upload_id, {"final_rows": normalized})
    session_store.append_flags(upload_id, new_flags)
    session_store.session_mark_stage(upload_id, "normalization")
    # NOTE: stage_complete is NOT dispatched here — this is a merged endpoint.
    # The HTTP response (with diff_data) is the authoritative completion signal for the frontend.

    summary = {
        "year_flags": sum(1 for f in new_flags if "year" in f.get("issue", "")),
        "story_flags": sum(1 for f in new_flags if "story" in f.get("issue", "") or "stories" in f.get("issue", "")),
        "area_flags": sum(1 for f in new_flags if "area" in f.get("issue", "")),
        "value_flags": sum(1 for f in new_flags if "value" in f.get("issue", "")),
        "currency_flags": sum(1 for f in new_flags if "currency" in f.get("issue", "")),
    }

    # ── Build AI Summary ───────────────────────────────────────────────────
    total_rows = len(normalized)

    def _count_non_null(rows, key):
        return sum(1 for r in rows if r.get(key) is not None and str(r.get(key, "")).strip() != "")

    def _count_changed(before_rows, after_rows, key):
        changed = 0
        for b, a in zip(before_rows, after_rows):
            bv = str(b.get(key, "") or "").strip()
            av = str(a.get(key, "") or "").strip()
            if bv != av and (bv or av):
                changed += 1
        return changed

    if target_format == "AIR":
        year_key, retro_key = "YearBuilt", "YearRetrofitted"
        stories_key, bldg_key = "NumberOfStories", "RiskCount"
        area_key = "GrossArea"
        val_keys = ["BuildingValue", "ContentsValue", "TimeElementValue"]
        currency_key = "Currency"
        sprinkler_key = "SprinklerSystem"
        roof_key, wall_key, found_key, soft_key = "RoofGeometry", "WallSiding", "FoundationType", "SoftStory"
    else:
        year_key, retro_key = "YEARBUILT", "YEARUPGRAD"
        stories_key, bldg_key = "NUMSTORIES", "NUMBLDGS"
        area_key = "FLOORAREA"
        val_keys = ["EQCV1VAL", "EQCV2VAL", "EQCV3VAL"]
        currency_key = "EQCV1LCUR"
        sprinkler_key = "SPRINKLER"
        roof_key, wall_key, found_key, soft_key = "ROOFGEOM", "CLADDING", "FOUNDATION", "SOFTSTORY"

    year_filled = _count_non_null(normalized, year_key)
    year_changed = _count_changed(final_rows_canonical, normalized, year_key)
    stories_filled = _count_non_null(normalized, stories_key)
    stories_changed = _count_changed(final_rows_canonical, normalized, stories_key)
    bldg_filled = _count_non_null(normalized, bldg_key)
    area_filled = _count_non_null(normalized, area_key)
    area_changed = _count_changed(final_rows_canonical, normalized, area_key)
    cur_filled = _count_non_null(normalized, currency_key)

    # ── Call LLM for narrative summary ─────────────────────────────────────────
    norm_stats_context = {
        "step": "normalization",
        "target_format": target_format,
        "total_rows": total_rows,
        "total_flags": len(new_flags),
        "year_stats": {"filled": year_filled, "changed": year_changed, "flags": summary["year_flags"]},
        "stories_stats": {"filled": stories_filled, "changed": stories_changed, "flags": summary["story_flags"]},
        "building_count_filled": bldg_filled,
        "area_stats": {"filled": area_filled, "changed": area_changed, "flags": summary["area_flags"]},
        "value_stats": {vk: {"filled": _count_non_null(normalized, vk), "changed": _count_changed(final_rows_canonical, normalized, vk)} for vk in val_keys},
        "currency_stats": {"filled": cur_filled, "flags": summary["currency_flags"]},
        "modifier_stats": {},
        "sample_flags": [{"field": f.get("field", ""), "issue": f.get("issue", ""), "original": str(f.get("original_value", ""))} for f in new_flags[:6]],
    }
    # Collect modifier stats
    for label, key in [("sprinkler", sprinkler_key), ("roof", roof_key), ("wall", wall_key),
                        ("foundation", found_key), ("soft_story", soft_key)]:
        filled = _count_non_null(normalized, key)
        changed = _count_changed(final_rows_canonical, normalized, key)
        if filled or changed:
            norm_stats_context["modifier_stats"][label] = {"filled": filled, "changed": changed}

    dispatch_event(upload_id, agent_id, "log", "Generating AI summary of normalization outcomes…")
    summary_text = _generate_llm_summary("normalization", norm_stats_context, "Values successfully normalized.")

    sample_rows = normalized[:10]
    headers_out = list(normalized[0].keys()) if normalized else []

    result = {
        "total_rows": len(normalized),
        "flags_added": len(new_flags),
        "sample": sample_rows,
        "headers": headers_out,
        "normalization_summary": summary
    }

    dispatch_event(upload_id, agent_id, "done", result=result)

    diff_data = _get_session_diff_data(session, "normalize")

    return NormalizeResponse(upload_id=upload_id, summary_text=summary_text, diff_data=diff_data, **result)


# â”€â”€ Slip Summary â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def _safe_float(v) -> float:
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().replace(",", "")
    for sym in ["$", "£", "€", "¥", "₹"]:
        s = s.replace(sym, "")
    try: return float(s)
    except (ValueError, TypeError): return 0.0

def _bucket_year(year_val) -> str:
    s = str(year_val or "").strip()
    try:
        y = int(float(s))
        if y <= 0: return "Unknown"
        if y < 1995: return "Pre 1995"
        if y <= 2001: return "1995 – 2001"
        if y <= 2010: return "2002 – 2010"
        if y <= 2017: return "2011 – 2017"
        return "Post 2017"
    except (ValueError, TypeError): return "Unknown"

def _bucket_stories(stories_val) -> str:
    s = str(stories_val or "").strip()
    try:
        n = int(float(s))
        if n <= 0: return "Unknown"
        if n == 1: return "1"
        if n <= 3: return "2–3"
        if n <= 7: return "4–7"
        return "7+"
    except (ValueError, TypeError): return "Unknown"

@app.get("/summary/{upload_id}", tags=["Output"])
def slip_summary(upload_id: str):
    session = _get_session_or_404(upload_id)

    target = session.get("target_format", "AIR")

    # Prefer the most-normalized rows available:
    # final_rows (CatAI full pipeline) → geo_rows (geocoded + address-normalized) → raw_rows
    final_rows = session.get("final_rows") or []
    if not final_rows:
        final_rows = session.get("geo_rows") or []
    if not final_rows:
        final_rows = session.get("raw_rows") or []

    if not final_rows:
        raise HTTPException(status_code=422, detail="No processed rows available. Run the pipeline first.")

    if target == "AIR":
        bldg_col, cont_col, bi_col, tiv_col = "BuildingValue", "ContentsValue", "TimeElementValue", "TIV"
        occ_col, const_col = "Occupancy_Description", "Construction_Description"
        year_col, stories_col = "YearBuilt", "NumberOfStories"
        country_col, state_col, city_col, street_col, zip_col = "CountryISO", "Area", "City", "Street", "PostalCode"
        loc_id_col = "LocationID"
    else:
        bldg_col, cont_col, bi_col, tiv_col = "EQCV1VAL", "EQCV2VAL", "EQCV3VAL", "TIV"
        occ_col, const_col = "Occupancy_Description", "Construction_Description"
        year_col, stories_col = "YEARBUILT", "NUMSTORIES"
        country_col, state_col, city_col, street_col, zip_col = "CNTRYCODE", "STATECODE", "CITY", "STREETNAME", "POSTALCODE"
        loc_id_col = "LOCNUM"

    # If geo_rows are being used (not yet CatAI mapped), dynamically resolve column names
    # from actual keys in the first row so we don't silently miss all data.
    if final_rows:
        sample_keys = set(final_rows[0].keys())

        def _resolve(preferred, *fallbacks):
            if preferred in sample_keys:
                return preferred
            for fb in fallbacks:
                if fb in sample_keys:
                    return fb
            return preferred   # keep original so result is just empty, not an error

        bldg_col     = _resolve(bldg_col, "BuildingValue", "BLDG_VALUE", "BldgValue", "Floor Area (sqft)")
        cont_col     = _resolve(cont_col, "ContentsValue", "CONT_VALUE", "ContValue")
        bi_col       = _resolve(bi_col, "TimeElementValue", "BI_VALUE", "BIValue")
        tiv_col      = _resolve(tiv_col, "TIV", "TotalInsuredValue", "Total Value")
        year_col     = _resolve(year_col, "YearBuilt", "YEARBUILT", "Year Built", "year_built")
        stories_col  = _resolve(stories_col, "NumberOfStories", "NUMSTORIES", "Floors", "Stories", "num_stories")
        country_col  = _resolve(country_col, "CountryISO", "CNTRYCODE", "Country")
        state_col    = _resolve(state_col, "Area", "STATECODE", "State", "Province")
        city_col     = _resolve(city_col, "City", "CITY", "Town")
        street_col   = _resolve(street_col, "Street", "STREETNAME", "Address", "_CombinedAddress")
        zip_col      = _resolve(zip_col, "PostalCode", "POSTALCODE", "Zip", "ZipCode", "Postal")

    total_bldg = total_cont = total_bi = total_tiv_col = 0.0
    country_state_map: Dict[str, Dict] = {}
    loc_rows = []
    occ_map: Dict[str, float] = {}
    const_map: Dict[str, float] = {}
    year_map: Dict[str, float] = {}
    story_map: Dict[str, float] = {}

    for row in final_rows:
        bldg, cont, bi = _safe_float(row.get(bldg_col)), _safe_float(row.get(cont_col)), _safe_float(row.get(bi_col))
        raw_tiv = _safe_float(row.get(tiv_col))

        if raw_tiv > 0:
            row_tiv = raw_tiv
            total_bldg += bldg; total_cont += cont; total_bi += bi; total_tiv_col += raw_tiv
        else:
            row_tiv = bldg + cont + bi
            total_bldg += bldg; total_cont += cont; total_bi += bi

        country = str(row.get(country_col) or "Unknown").strip() or "Unknown"
        state   = str(row.get(state_col) or "NA").strip() or "NA"
        cs_key  = f"{country}||{state}"
        if cs_key not in country_state_map: country_state_map[cs_key] = {"country": country, "state": state, "count": 0, "tiv": 0.0}
        country_state_map[cs_key]["count"] += 1
        country_state_map[cs_key]["tiv"]   += row_tiv

        loc_rows.append({
            "loc_id":  str(row.get(loc_id_col) or ""), "address": str(row.get(street_col) or ""),
            "city":    str(row.get(city_col) or ""), "state":   str(row.get(state_col) or ""),
            "zip":     str(row.get(zip_col) or ""), "tiv":     row_tiv,
        })

        occ = str(row.get(occ_col) or row.get("OccupancyCode") or "Unknown").strip() or "Unknown"
        occ_map[occ] = occ_map.get(occ, 0.0) + row_tiv

        const = str(row.get(const_col) or row.get("ConstructionCode") or "Unknown").strip() or "Unknown"
        const_map[const] = const_map.get(const, 0.0) + row_tiv

        yb = _bucket_year(row.get(year_col))
        year_map[yb] = year_map.get(yb, 0.0) + row_tiv

        st = _bucket_stories(row.get(stories_col))
        story_map[st] = story_map.get(st, 0.0) + row_tiv

    grand_total = total_tiv_col if total_tiv_col > 0 else (total_bldg + total_cont + total_bi)
    top_locs = sorted(loc_rows, key=lambda r: r["tiv"], reverse=True)[:10]
    cs_list = sorted(country_state_map.values(), key=lambda r: r["tiv"], reverse=True)

    year_order = ["Unknown", "Pre 1995", "1995 – 2001", "2002 – 2010", "2011 – 2017", "Post 2017"]
    year_dist  = [{"label": k, "tiv": year_map.get(k, 0.0)} for k in year_order if k in year_map]

    story_order = ["Unknown", "1", "2–3", "4–7", "7+"]
    story_dist  = [{"label": k, "tiv": story_map.get(k, 0.0)} for k in story_order if k in story_map]

    return {
        "total_risks": len(final_rows),
        "location_values": {"building": total_bldg, "contents": total_cont, "bi": total_bi, "total": grand_total},
        "country_state": cs_list,
        "top_locations": top_locs,
        "occupancy_dist": [{"label": k, "tiv": v} for k, v in sorted(occ_map.items(), key=lambda x: -x[1])],
        "construction_dist": [{"label": k, "tiv": v} for k, v in sorted(const_map.items(), key=lambda x: -x[1])],
        "year_built_dist": year_dist,
        "stories_dist": story_dist,
    }


# â”€â”€ Download â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.get("/download/{upload_id}", tags=["Output"])
def download(upload_id: str, format: str = Query("xlsx", pattern="^(xlsx|txt)$")):
    session = _get_session_or_404(upload_id)
    _require_stage(session, "normalization", "/download")

    final_rows = session.get("final_rows", [])
    unmapped_cols = session.get("unmapped_cols", [])
    flags = session.get("flags", [])
    target = session.get("target_format", "AIR")
    short_id = upload_id[:8]

    if format == "txt":
        buf = build_tsv(final_rows, unmapped_cols, target)
        filename = f"cat_output_{short_id}.txt"
        media_type = "text/tab-separated-values"
    else:
        buf = build_xlsx(final_rows, unmapped_cols, flags, target, upload_id)
        filename = f"cat_output_{short_id}.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return StreamingResponse(buf, media_type=media_type, headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.get("/download-account/{upload_id}", tags=["Output"])
def download_account(upload_id: str, format: str = Query("xlsx", pattern="^(xlsx|txt)$")):
    """Download the account file output as XLSX or TXT."""
    session = _get_session_or_404(upload_id)
    _require_stage(session, "normalization", "/download-account")

    final_rows = session.get("final_rows", [])
    target = session.get("target_format", "AIR")
    short_id = upload_id[:8]

    if format == "txt":
        buf = build_account_tsv(final_rows, target)
        filename = f"account_output_{short_id}.txt"
        media_type = "text/tab-separated-values"
    else:
        buf = build_account_xlsx(final_rows, target)
        filename = f"account_output_{short_id}.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    logger.info(f"Session {upload_id}: account download requested ({format}), {len(final_rows)} source rows")
    return StreamingResponse(buf, media_type=media_type, headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.get("/preview-location/{upload_id}", tags=["Output"])
def preview_location(upload_id: str):
    """Return an exact 5-row preview of the generated location file."""
    session = _get_session_or_404(upload_id)
    _require_stage(session, "normalization", "/preview-location")
    
    final_rows = session.get("final_rows", [])
    target = session.get("target_format", "AIR")
    
    # We copy the first 5 rows because generate_location_data mutates the dicts (e.g. adding RMS fields)
    sample_rows = [dict(r) for r in final_rows[:5]]
    headers, formatted_rows = generate_location_data(sample_rows, target)
    
    return {"headers": headers, "sample": formatted_rows}

@app.get("/preview-account/{upload_id}", tags=["Output"])
def preview_account(upload_id: str):
    """Return an exact preview of the generated account file."""
    session = _get_session_or_404(upload_id)
    _require_stage(session, "normalization", "/preview-account")
    
    final_rows = session.get("final_rows", [])
    target = session.get("target_format", "AIR")
    
    # Generate the account file rows (it already aggregates the whole dataset)
    headers, acc_rows = _get_account_rows(final_rows, target)
    
    return {"headers": headers, "sample": acc_rows[:5]}


# ── Pipeline Diff ──────────────────────────────────────────────────────────────────

def _get_session_diff_data(session: dict, step: str) -> dict:
    """
    Return before/after table data for a specific pipeline step, capped at 100 rows.
    Includes all columns that appear in the final Excel output.
    Also returns a `pairs` list: [{before_col, after_col, label}] so the UI can
    render old → new columns adjacently.
    """
    target = session.get("target_format", "AIR")

    column_map = session.get("column_map", {})
    raw_rows = session.get("raw_rows", [])

    # Reverse map: canonical → source column name
    canonical_to_src: Dict[str, str] = {}
    
    if step == "map-codes":
        _require_stage(session, "code_mapping", f"/session-diff (step={step})")
        for src, can in column_map.items():
            if can and can not in canonical_to_src:
                canonical_to_src[can] = src
    elif step == "normalize":
        _require_stage(session, "normalization", f"/session-diff (step={step})")
        for src, can in column_map.items():
            if can and can not in canonical_to_src:
                canonical_to_src[can] = src

    def get_source_cols(canonicals: set) -> List[str]:
        return [src for src, can in column_map.items() if can in canonicals]

    # pairs = [{"label": str, "before": src_col | None, "after": canonical_col | None}]
    pairs: List[Dict] = []
    after_rows: List[Dict] = []

    if step == "geocode":
        _require_stage(session, "geocoding", f"/session-diff (step={step})")
        after_rows = session.get("geo_rows", [])
        # Use normalized rows (pre-geocode) as the before-state
        raw_rows = session.get("raw_rows", [])

        # In multiagent, Address Normalization runs BEFORE geocoding and generates "_CombinedAddress".
        headers = session.get("headers", [])
        full_addr_src = "_CombinedAddress" if "_CombinedAddress" in headers else None
        
        # Fallback to other address fields if somehow _CombinedAddress is absent
        if not full_addr_src:
            for h in headers:
                if "address" in h.lower() or "street" in h.lower():
                    full_addr_src = h
                    break
                    
        full_address_mode = bool(full_addr_src)

        if target == "AIR":
            addr_fields = [
                ("Street",      "Street"),
                ("City",        "City"),
                ("Area",        "Area"),
                ("PostalCode",  "PostalCode"),
                ("CountryISO",  "CountryISO"),
                ("Latitude",    "Latitude"),
                ("Longitude",   "Longitude"),
            ]
        else:
            addr_fields = [
                ("STREETNAME",  "STREETNAME"),
                ("CITY",        "CITY"),
                ("STATECODE",   "STATECODE"),
                ("POSTALCODE",  "POSTALCODE"),
                ("CNTRYCODE",   "CNTRYCODE"),
                ("Latitude",    "Latitude"),
                ("Longitude",   "Longitude"),
            ]

        if full_address_mode:
            # Single input → many extracted outputs.
            for canonical, after_col in addr_fields:
                pairs.append({
                    "label": after_col,
                    "before": full_addr_src,    # always the same source column
                    "after":  after_col,
                    "before_is_full_address": True,  # hint for the UI
                })
        else:
            # Fallback if no initial address was provided
            for canonical, after_col in addr_fields:
                pairs.append({"label": canonical, "before": None, "after": after_col})

        # Geocoding-only outputs (no source equivalent)
        pairs.append({"label": "GeocodingStatus", "before": None, "after": "GeocodingStatus"})
        pairs.append({"label": "Geosource",       "before": None, "after": "Geosource"})

    elif step == "map-codes":
        _require_stage(session, "code_mapping", f"/session-diff (step={step})")
        after_rows = session.get("final_rows", session.get("geo_rows", []))

        if target == "AIR":
            code_pairs = [
                ("OccupancyCodeType",    "OccupancyCodeType",    "Occupancy_Code",         "Occ Code"),
                ("OccupancyCode",        "OccupancyCode",        "Occupancy_Description",  "Occ Description"),
                (None,                   None,                   "Occupancy_Method",       "Occ Method"),
                ("ConstructionCodeType", "ConstructionCodeType", "Construction_Code",      "Const Code"),
                ("ConstructionCode",     "ConstructionCode",     "Construction_Description","Const Description"),
                (None,                   None,                   "Construction_Method",    "Const Method"),
            ]
        else:
            code_pairs = [
                ("OCCSCHEME",  "OCCSCHEME",  "Occupancy_Code",         "Occ Code"),
                ("OCCTYPE",    "OCCTYPE",    "Occupancy_Description",  "Occ Description"),
                (None,         None,         "Occupancy_Method",       "Occ Method"),
                ("BLDGSCHEME", "BLDGSCHEME", "Construction_Code",      "Const Code"),
                ("BLDGCLASS",  "BLDGCLASS",  "Construction_Description","Const Description"),
                (None,         None,         "Construction_Method",    "Const Method"),
            ]

        for canonical, _, after_col, label in code_pairs:
            src = canonical_to_src.get(canonical) if canonical else None
            pairs.append({"label": label, "before": src, "after": after_col})

    elif step == "normalize":
        _require_stage(session, "normalization", f"/session-diff (step={step})")
        after_rows = session.get("final_rows", [])

        if target == "AIR":
            norm_pairs = [
                ("YearBuilt",        "YearBuilt"),
                ("YearRetrofitted",  "YearRetrofitted"),
                ("NumberOfStories",  "NumberOfStories"),
                ("RiskCount",        "RiskCount"),
                ("GrossArea",        "GrossArea"),
                ("BuildingValue",    "BuildingValue"),
                ("ContentsValue",    "ContentsValue"),
                ("TimeElementValue", "TimeElementValue"),
                ("Currency",         "Currency"),
                ("LineOfBusiness",   "LineOfBusiness"),
                ("SprinklerSystem",  "SprinklerSystem"),
                ("RoofGeometry",     "RoofGeometry"),
                ("FoundationType",   "FoundationType"),
                ("WallSiding",       "WallSiding"),
                ("WallType",         "WallType"),
                ("SoftStory",        "SoftStory"),
            ]
        else:
            norm_pairs = [
                ("YEARBUILT",   "YEARBUILT"),
                ("YEARUPGRAD",  "YEARUPGRAD"),
                ("NUMSTORIES",  "NUMSTORIES"),
                ("NUMBLDGS",    "NUMBLDGS"),
                ("FLOORAREA",   "FLOORAREA"),
                ("EQCV1VAL",    "EQCV1VAL"),
                ("EQCV2VAL",    "EQCV2VAL"),
                ("EQCV3VAL",    "EQCV3VAL"),
                ("EQCV1LCUR",   "EQCV1LCUR"),
                ("SPRINKLER",   "SPRINKLER"),
                ("ROOFGEOM",    "ROOFGEOM"),
                ("FOUNDATION",  "FOUNDATION"),
                ("CLADDING",    "CLADDING"),
                ("WALLTYPE",    "WALLTYPE"),
                ("SOFTSTORY",   "SOFTSTORY"),
            ]

        for canonical, after_col in norm_pairs:
            src = canonical_to_src.get(canonical)
            pairs.append({"label": canonical, "before": src, "after": after_col})

    else:
        raise HTTPException(status_code=400, detail="Invalid step")

    # Derive flat before/after column lists from pairs (deduplicated, preserving order)
    seen_before: set = set()
    before_cols = []
    for p in pairs:
        c = p.get("before")
        if c and c not in seen_before:
            seen_before.add(c)
            before_cols.append(c)
    after_cols = [p["after"] for p in pairs if p.get("after")]

    # Detect full_address_mode from pairs metadata
    full_address_mode = any(p.get("before_is_full_address") for p in pairs)
    full_address_src  = next((p["before"] for p in pairs if p.get("before_is_full_address")), None)

    limit = 100
    rows_data = []

    for i in range(min(len(raw_rows), len(after_rows))):
        before_data = {c: raw_rows[i].get(c) for c in before_cols}
        after_data  = {c: after_rows[i].get(c)  for c in after_cols}
        # Only include row if it has some relevant data on either side
        if any(v is not None and str(v).strip() != "" for v in list(before_data.values()) + list(after_data.values())):
            rows_data.append({"before": before_data, "after": after_data})

    return {
        "step": step,
        "columns": {"before": before_cols, "after": after_cols},
        "pairs": pairs,
        "full_address_mode": full_address_mode,
        "full_address_src":  full_address_src,
        "rows": rows_data[:limit],
        "total": len(rows_data)
    }

@app.get("/session-diff/{upload_id}", tags=["Pipeline"])
def session_diff(upload_id: str, step: str = Query(..., pattern="^(geocode|map-codes|normalize|normalize-address)$")):
    session = _get_session_or_404(upload_id)
    return _get_session_diff_data(session, step)


# ── Entry point ───────────────────────────────────────────────────────────

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
